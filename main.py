import os
from datetime import datetime
from typing import List, Dict
import pandas as pd
import re
from langchain_community.document_loaders import PyMuPDFLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from openai import OpenAI
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils.dataframe import dataframe_to_rows

# Qwen - Plus model configuration
# Maximum input length for the Qwen - Plus model
MAX_INPUT_LENGTH_PLUS = 98304
# Maximum input length for the Qwen - Turbo model
MAX_INPUT_LENGTH_TURBO = 1000000
# Maximum length of the reply
MAX_REPLY_LENGTH = 8192

# Column names for the Excel output
COLUMN_NAMES = {
    'A': 'Publication Date',
    'B': 'Author Names',
    'C': 'Journal Name',
    'D': 'Article Title',
    'E': 'Keywords',
    'F': 'Summary',
    'G': 'Core Findings',
    'H': 'Variables',
    'I': 'Theory Name',
    'J': 'Theoretical Framework',
    'K': 'Methodology',
    'L': 'Red Flags',
    'M': 'Relevance to Research'
}

# Words that should be displayed in bold in the Excel output
BOLD_WORDS = [
    "Rating",
]


class PDFProcessor:
    @staticmethod
    def process_folder(folder_path: str) -> List[str]:
        """
        Process a folder and return a list of paths to all PDF files in it.
        :param folder_path: The path of the folder to be processed.
        :return: A list containing the paths of all PDF files.
        """
        return [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.lower().endswith('.pdf')]

    @staticmethod
    def read_pdf(file_path: str) -> str:
        """
        Read the content of a PDF file.
        :param file_path: The path of the PDF file to be read.
        :return: The text content of the PDF file.
        :raises RuntimeError: If an error occurs while reading the PDF file.
        """
        try:
            loader = PyMuPDFLoader(file_path)
            documents = loader.load()
            text_splitter = RecursiveCharacterTextSplitter(
                chunk_size=MAX_INPUT_LENGTH_PLUS - 4000,
                chunk_overlap=0,
                length_function=len
            )
            chunks = text_splitter.split_documents(documents)
            text = "\n".join([chunk.page_content for chunk in chunks])
            print(f"Successfully read PDF file {file_path}, text length: {len(text)}")
            return text
        except Exception as e:
            raise RuntimeError(f"Failed to read PDF: {str(e)}")


class QwenAPI:
    def __init__(self, api_key: str, user_prompt_path):
        """
        Initialize the Qwen API client.
        :param api_key: The API key for accessing the Qwen API.
        :param user_prompt_path: The path to the user prompt template file.
        """
        self.client = OpenAI(
            api_key=api_key,
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"
        )
        self.user_prompt_template = self._read_prompt(user_prompt_path)

    def _read_prompt(self, path):
        """
        Read the user prompt template from a file.
        :param path: The path of the file containing the user prompt template.
        :return: The content of the user prompt template.
        """
        with open(path, 'r', encoding='utf-8') as file:
            return file.read()

    def get_summary(self, text: str, model_name) -> List[Dict]:
        """
        Get a summary of the given text using the specified Qwen model.
        :param text: The text to be summarized.
        :param model_name: The name of the Qwen model to use (e.g., "qwen-plus", "qwen-turbo").
        :return: A list of dictionaries containing the summary result or an error message.
        """
        max_length = MAX_INPUT_LENGTH_PLUS if model_name == "qwen-plus" else MAX_INPUT_LENGTH_TURBO if model_name == "qwen-turbo" else None
        if not max_length:
            print(f"Unsupported model: {model_name}")
            return [{"error": f"Unsupported model: {model_name}"}]

        total_length = len(text) + len(self.user_prompt_template)
        if total_length > max_length:
            if model_name == "qwen-plus":
                print(f"Warning: Input length {total_length} exceeds Qwen - Plus limit {max_length}, switching to Qwen - Turbo.")
                return self.get_summary(text, "qwen-turbo")
            else:
                print(f"Warning: Input length {total_length} exceeds Qwen - Turbo limit {max_length}.")
                return [{"error": f"Input length exceeds model limit {max_length}"}]

        def clean_text(text):
            """
            Clean the text by removing extra whitespace.
            :param text: The text to be cleaned.
            :return: The cleaned text.
            """
            return re.sub(r'\s+', ' ', text).strip()

        user_prompt = self.user_prompt_template.format(text=clean_text(text))
        messages = [{"role": "user", "content": user_prompt}]
        print(f"Sending request to {model_name}, request content:\n{messages}")

        try:
            completion = self.client.chat.completions.create(
                model=model_name,
                messages=messages,
                stream=False
            )
            answer_content = completion.choices[0].message.content
            print("=" * 20 + f"Complete Reply from {model_name}" + "=" * 20 + "\n")
            print(answer_content)
            return [{"content": answer_content}]
        except Exception as e:
            print(f"API request failed: {str(e)}")
            return [{"error": f"API request failed: {str(e)}"}]


class ResultExporter:
    @staticmethod
    def to_excel(data: List[Dict], output_path: str = "research_summary.xlsx"):
        """
        Export the processing results to an Excel file.
        :param data: A list of dictionaries containing the processing results.
        :param output_path: The path of the output Excel file.
        :return: The absolute path of the saved Excel file.
        """
        current_time = datetime.now().strftime("%Y%m%d%H%M%S")
        new_file_name = f"{os.path.splitext(output_path)[0]}_{current_time}{os.path.splitext(output_path)[1]}"

        rows = []
        for result in data:
            file_path = result['file_path']
            process_time = result['process_time']
            for response in result['raw_responses']:
                row = {'File Path': file_path, 'Process Time': process_time}
                if 'error' in response:
                    row['Error'] = response['error']
                    rows.append(row)
                    continue

                content = response['content']
                sections = ResultExporter.parse_sections(content)
                for key, column in COLUMN_NAMES.items():
                    raw_value = sections.get(key, '')
                    cleaned_value = ResultExporter.clean_text(raw_value)
                    row[column] = cleaned_value
                rows.append(row)

        columns = ['File Path', 'Process Time'] + list(COLUMN_NAMES.values())
        df = pd.DataFrame(rows, columns=columns)
        df['Publication Date'] = df['Publication Date'].apply(ResultExporter.format_date)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'

        # Set the font of the table header to bold
        header_font = Font(name='Times New Roman', bold=True)
        for cell in ws[1]:
            cell.font = header_font

        # Write the DataFrame to the worksheet using dataframe_to_rows
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                if isinstance(value, str):
                    rich_text = CellRichText()
                    pos = 0
                    bold_font = InlineFont(rFont='Times New Roman', b=True)
                    normal_font = InlineFont(rFont='Times New Roman')

                    # Make the text between numbers and colons bold
                    while pos < len(value):
                        match = re.search(r'(\d+\.\s*)(.*?)(:)', value[pos:])
                        if match:
                            start = match.start() + pos
                            end = match.end() + pos
                            if start > pos:
                                rich_text.append(TextBlock(normal_font, value[pos:start]))
                            rich_text.append(TextBlock(bold_font, match.group(1) + match.group(2) + match.group(3)))
                            pos = end
                        else:
                            # Process the specified BOLD_WORDS
                            found = False
                            for word in BOLD_WORDS:
                                index = value.find(word, pos)
                                if index != -1:
                                    if index > pos:
                                        rich_text.append(TextBlock(normal_font, value[pos:index]))
                                    rich_text.append(TextBlock(bold_font, word))
                                    pos = index + len(word)
                                    found = True
                                    break
                            if not found:
                                rich_text.append(TextBlock(normal_font, value[pos:]))
                                break
                    ws.cell(row=r_idx, column=c_idx, value=rich_text)
                else:
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.font = Font(name='Times New Roman')

        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = 18

        alignment = Alignment(wrap_text=True, vertical='top')
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment

        for idx, row in enumerate(ws.iter_rows(), start=1):
            ws.row_dimensions[row[0].row].height = 30 if idx == 1 else 200

        wb.save(new_file_name)
        print(f"Results saved to Excel file: {os.path.abspath(new_file_name)}")
        return os.path.abspath(new_file_name)

    @staticmethod
    def parse_sections(content: str) -> Dict[str, str]:
        """
        Parse the content into sections.
        :param content: The content to be parsed.
        :return: A dictionary containing the parsed sections.
        """
        sections = {}
        lines = content.split('\n')
        current_section = None
        current_value = []

        for line in lines:
            stripped = line.strip()
            if not stripped:
                continue

            match = re.match(r'^([A-Z])\.', stripped)
            if match:
                if current_section is not None:
                    sections[current_section] = '\n'.join(current_value).strip()
                current_section = match.group(1)
                content_part = re.sub(r'^[A-Z]\.\s*', '', stripped)
                current_value = [content_part]
            else:
                if current_section is not None:
                    current_value.append(stripped)

        if current_section is not None:
            sections[current_section] = '\n'.join(current_value).strip()

        valid_keys = set(COLUMN_NAMES.keys())
        return {key: value for key, value in sections.items() if key in valid_keys}

    @staticmethod
    def clean_text(text: str) -> str:
        """
        Clean the text by removing Markdown formatting and special characters.
        :param text: The text to be cleaned.
        :return: The cleaned text.
        """
        text = re.sub(r'###\s*', '', text)
        text = re.sub(r'^\s*[-]', '', text, flags=re.MULTILINE)
        return re.sub(r'\s+', ' ', text).strip()

    @staticmethod
    def format_date(date_str):
        """
        Format the date string.
        :param date_str: The date string to be formatted.
        :return: The formatted date string.
        """
        if re.match(r'^\d{4}$', date_str):
            return date_str
        try:
            date_obj = pd.to_datetime(date_str, errors='coerce')
            return date_obj.strftime('%Y/%m/%d') if pd.notna(date_obj) else date_str
        except:
            return date_str


def main():
    """
    The main function to execute the entire process, including reading PDFs, getting summaries, and exporting results to Excel.
    """
    CONFIG = {
        "api_key": "Add you API KEY here",
        "pdf_folder": r"C:\Users\xinjiehuang\Desktop\research\trust\test",
        "output_file": "research_summary.xlsx",
        "user_prompt_path": r"C:\Users\xinjiehuang\Desktop\research\trust\prompts\user_prompt.txt"
    }

    processor = PDFProcessor()
    api_client = QwenAPI(CONFIG["api_key"], CONFIG["user_prompt_path"])
    results = []

    for pdf_path in processor.process_folder(CONFIG["pdf_folder"]):
        try:
            print(f"Processing: {os.path.basename(pdf_path)}")
            text = processor.read_pdf(pdf_path)
            raw_responses = api_client.get_summary(text, "qwen-plus")
            pdf_filename = os.path.basename(pdf_path)
            txt_path = os.path.join(os.path.dirname(pdf_path), "qwen_answers", f"{os.path.splitext(pdf_filename)[0]}.txt")
            os.makedirs(os.path.dirname(txt_path), exist_ok=True)
            with open(txt_path, "w", encoding="utf-8") as f:
                f.write(raw_responses[0]["content"])
            results.append({
                "file_path": pdf_path,
                "process_time": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "raw_responses": raw_responses
            })
        except Exception as e:
            print(f"Processing failed for {pdf_path}: {str(e)}")

    output_path = ResultExporter.to_excel(results, CONFIG["output_file"])
    print(f"Processing completed! Results saved to: {output_path}")


if __name__ == "__main__":
    main()
